from __future__ import annotations

import json
import re
import sys
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import (
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


FOREST = colors.HexColor('#173B33')
FOREST_SOFT = colors.HexColor('#2D5B4E')
CREAM = colors.HexColor('#F6F2E9')
PAPER = colors.HexColor('#FFFEFA')
SAGE = colors.HexColor('#DFE8DF')
GOLD = colors.HexColor('#B89B61')
INK = colors.HexColor('#21342F')
MUTED = colors.HexColor('#6F7B76')
LINE = colors.HexColor('#D8DDD6')
STRIPE = colors.HexColor('#F3F4EF')
ABSENT = colors.HexColor('#FCE4D6')

BODY_FONT = 'HeiseiKakuGo-W5'
DISPLAY_FONT = 'HeiseiMin-W3'


def register_fonts() -> None:
    pdfmetrics.registerFont(UnicodeCIDFont(BODY_FONT))
    pdfmetrics.registerFont(UnicodeCIDFont(DISPLAY_FONT))


def clean_text(value: object) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def cell_text(sheet, row: int, column: int) -> str:
    return clean_text(sheet.cell(row=row, column=column).value)


def rgb_from_cell(cell, attribute: str) -> str | None:
    color = getattr(cell, attribute, None)
    if color is None or color.type != 'rgb' or not color.rgb:
        return None
    rgb = color.rgb[-6:]
    return rgb if re.fullmatch(r'[0-9A-Fa-f]{6}', rgb) else None


def teacher_label(sheet_name: str) -> str:
    cleaned = re.sub(r'講師$', '', sheet_name).strip()
    tokens = re.split(r'[\s\u3000]+', cleaned)
    surname = tokens[0] if len(tokens) > 1 else ''.join(list(cleaned)[:2])
    return f'{surname or cleaned}講師'


def detect_period(sheet) -> tuple[str, str]:
    for row in range(11, sheet.max_row + 1):
        match = re.search(r'(\d{4})[/-](\d{1,2})', cell_text(sheet, row, 6))
        if match:
            year = match.group(1)
            month = int(match.group(2))
            return f'{year}{month:02d}', f'{year}年 {month}月'
    month_match = re.search(r'(\d{1,2})月', cell_text(sheet, 2, 1))
    month = int(month_match.group(1)) if month_match else 1
    return f'2026{month:02d}', f'2026年 {month}月'


def number_value(value: object) -> float:
    return float(value) if isinstance(value, (int, float)) else 0.0


def metric_values(sheet) -> list[tuple[str, str]]:
    labels = ['1:2', '1:2 特能', '1:1 特能', '集団指導', '事務作業', '英会話']
    values = []
    for column in range(8, 14):
        total = sum(number_value(sheet.cell(row=row, column=column).value) for row in range(11, sheet.max_row + 1))
        values.append(f'{int(total)}分')
    return list(zip(labels, values))


def paragraph(text: str, style: ParagraphStyle) -> Paragraph:
    safe = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('\n', '<br/>')
    return Paragraph(safe or ' ', style)


def build_identity_block(sheet, teacher: str, period_label: str):
    small_caps = ParagraphStyle(
        'SmallCaps', fontName=BODY_FONT, fontSize=6.3, leading=7.2,
        textColor=GOLD, tracking=2, spaceAfter=2 * mm,
    )
    teacher_style = ParagraphStyle(
        'Teacher', fontName=DISPLAY_FONT, fontSize=23, leading=26,
        textColor=FOREST, tracking=2,
    )
    subtitle_style = ParagraphStyle(
        'Subtitle', fontName=BODY_FONT, fontSize=7.4, leading=9,
        textColor=MUTED, tracking=1.2, spaceBefore=2 * mm,
    )
    stat_label = ParagraphStyle('StatLabel', fontName=BODY_FONT, fontSize=5.8, leading=7, textColor=MUTED)
    stat_value = ParagraphStyle('StatValue', fontName=DISPLAY_FONT, fontSize=12, leading=14, textColor=FOREST, alignment=TA_RIGHT)

    working_days = cell_text(sheet, 5, 3) or '-'
    individual_count = sum(
        1 for row in range(11, sheet.max_row + 1)
        if any(number_value(sheet.cell(row=row, column=column).value) > 0 for column in (8, 9, 10))
    )
    stats = Table(
        [
            [paragraph('月間勤務日数', stat_label), paragraph('個別授業回数', stat_label)],
            [paragraph(working_days, stat_value), paragraph(f'{individual_count} 回', stat_value)],
        ],
        colWidths=[31 * mm, 31 * mm],
        rowHeights=[8 * mm, 9 * mm],
    )
    stats.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), PAPER),
        ('BOX', (0, 0), (-1, -1), 0.5, LINE),
        ('INNERGRID', (0, 0), (-1, -1), 0.4, LINE),
        ('LINEBEFORE', (0, 0), (0, -1), 2.2, GOLD),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3 * mm),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3 * mm),
        ('TOPPADDING', (0, 0), (-1, -1), 1.2 * mm),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1.2 * mm),
    ]))

    teacher_block = [
        paragraph('TEACHER', small_caps),
        paragraph(teacher, teacher_style),
        paragraph('勤務時間集計表', subtitle_style),
    ]
    identity = Table([[teacher_block, stats]], colWidths=[116 * mm, 62 * mm])
    identity.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    return identity


def build_metric_grid(sheet):
    label_style = ParagraphStyle('MetricLabel', fontName=BODY_FONT, fontSize=5.4, leading=6.5, textColor=MUTED, alignment=TA_LEFT)
    value_style = ParagraphStyle('MetricValue', fontName=DISPLAY_FONT, fontSize=11, leading=13, textColor=FOREST, alignment=TA_LEFT)
    cards = []
    for index, (label, value) in enumerate(metric_values(sheet)):
        cards.append([paragraph(label, label_style), paragraph(value, value_style), index == 0])
    table = Table(
        [[Table([[card[0]], [card[1]]], colWidths=[27.8 * mm], rowHeights=[7 * mm, 9 * mm]) for card in cards]],
        colWidths=[29.7 * mm] * 6,
    )
    commands = [
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX', (0, 0), (-1, -1), 0.45, LINE),
        ('INNERGRID', (0, 0), (-1, -1), 0.45, LINE),
        ('BACKGROUND', (0, 0), (0, 0), SAGE),
        ('BACKGROUND', (1, 0), (-1, 0), PAPER),
        ('LEFTPADDING', (0, 0), (-1, -1), 1 * mm),
        ('RIGHTPADDING', (0, 0), (-1, -1), 1 * mm),
        ('TOPPADDING', (0, 0), (-1, -1), 1.3 * mm),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1.3 * mm),
    ]
    table.setStyle(TableStyle(commands))
    return table


def build_detail_table(sheet):
    header_style = ParagraphStyle(
        'TableHeader', fontName=BODY_FONT, fontSize=4.3, leading=5.2,
        textColor=colors.white, alignment=TA_CENTER,
    )
    body_left = ParagraphStyle('BodyLeft', fontName=BODY_FONT, fontSize=4.2, leading=5.1, textColor=INK, alignment=TA_LEFT)
    body_center = ParagraphStyle('BodyCenter', parent=body_left, alignment=TA_CENTER)
    body_right = ParagraphStyle('BodyRight', parent=body_left, alignment=TA_RIGHT)
    headers = [paragraph(cell_text(sheet, 10, column), header_style) for column in range(1, 16)]
    rows = [headers]
    row_heights = [8.5 * mm]
    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), FOREST),
        ('BOX', (0, 0), (-1, -1), 0.45, LINE),
        ('INNERGRID', (0, 0), (-1, -1), 0.28, LINE),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0.6 * mm),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0.6 * mm),
        ('TOPPADDING', (0, 1), (-1, -1), 0.65 * mm),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 0.65 * mm),
    ]

    output_row = 1
    for source_row in range(11, sheet.max_row + 1):
        values = [cell_text(sheet, source_row, column) for column in range(1, 16)]
        if not any(values):
            rows.append([''] * 15)
            row_heights.append(2.8 * mm)
            style_commands.append(('BACKGROUND', (0, output_row), (-1, output_row), CREAM))
            output_row += 1
            continue
        rendered = []
        for column, value in enumerate(values, start=1):
            style = body_center if column in (4, 5, 6, 7) else body_right if column in (8, 9, 10, 11, 12, 13, 15) else body_left
            rendered.append(paragraph(value, style))
        rows.append(rendered)
        row_heights.append(None)
        style_commands.append(('BACKGROUND', (0, output_row), (-1, output_row), PAPER if output_row % 2 == 0 else STRIPE))

        name_fill = rgb_from_cell(sheet.cell(source_row, 1).fill, 'fgColor')
        if name_fill and name_fill.upper() == 'FCE4D6':
            style_commands.append(('BACKGROUND', (0, output_row), (0, output_row), ABSENT))
        for column in range(1, 16):
            font_rgb = rgb_from_cell(sheet.cell(source_row, column).font, 'color')
            if font_rgb and font_rgb.upper() not in {'000000', '21342F'}:
                style_commands.append(('TEXTCOLOR', (column - 1, output_row), (column - 1, output_row), colors.HexColor(f'#{font_rgb}')))
        output_row += 1

    width_weights = [92, 92, 78, 42, 42, 106, 106, 44, 54, 54, 60, 60, 60, 116, 50]
    total_width = 194 * mm
    total_weight = sum(width_weights)
    column_widths = [total_width * weight / total_weight for weight in width_weights]
    table = Table(rows, colWidths=column_widths, rowHeights=row_heights, repeatRows=1, splitByRow=1)
    table.setStyle(TableStyle(style_commands))
    return table


def draw_page(canvas, doc, teacher: str, period_label: str) -> None:
    width, height = A4
    canvas.saveState()
    canvas.setFillColor(CREAM)
    canvas.rect(0, 0, width, height, fill=1, stroke=0)
    canvas.setFillColor(FOREST)
    canvas.rect(0, height - 5 * mm, width * 0.73, 5 * mm, fill=1, stroke=0)
    canvas.setFillColor(GOLD)
    canvas.rect(width * 0.73, height - 5 * mm, width * 0.27, 5 * mm, fill=1, stroke=0)
    canvas.setFillColor(FOREST)
    canvas.setFont(DISPLAY_FONT, 11)
    canvas.drawString(8 * mm, height - 13 * mm, 'Re:Act')
    canvas.setFillColor(GOLD)
    canvas.setFont(BODY_FONT, 5.4)
    canvas.drawString(27 * mm, height - 12.7 * mm, 'WORK SUMMARY')
    canvas.setFillColor(PAPER)
    canvas.roundRect(width - 43 * mm, height - 17 * mm, 35 * mm, 8 * mm, 4 * mm, fill=1, stroke=0)
    canvas.setStrokeColor(LINE)
    canvas.roundRect(width - 43 * mm, height - 17 * mm, 35 * mm, 8 * mm, 4 * mm, fill=0, stroke=1)
    canvas.setFillColor(FOREST)
    canvas.setFont(BODY_FONT, 6.6)
    canvas.drawCentredString(width - 25.5 * mm, height - 14.3 * mm, period_label)
    canvas.setStrokeColor(GOLD)
    canvas.setLineWidth(0.35)
    canvas.line(8 * mm, 7 * mm, width - 8 * mm, 7 * mm)
    canvas.setFillColor(FOREST_SOFT)
    canvas.setFont(BODY_FONT, 5.2)
    canvas.drawString(8 * mm, 3.7 * mm, f'RE:ACT / WORK SUMMARY / {teacher}')
    canvas.drawRightString(width - 8 * mm, 3.7 * mm, f'{canvas.getPageNumber()}')
    canvas.restoreState()


def build_teacher_pdf(sheet, output_path: Path) -> tuple[str, str]:
    period_prefix, period_label = detect_period(sheet)
    label = teacher_label(sheet.title)
    doc = SimpleDocTemplate(
        str(output_path), pagesize=A4,
        leftMargin=8 * mm, rightMargin=8 * mm,
        topMargin=22 * mm, bottomMargin=11 * mm,
        title=f'{period_label} 勤務時間集計表 {label}',
        author='勤務時間集計ツール Re:Act',
    )
    detail_heading_style = ParagraphStyle(
        'DetailHeading', fontName=BODY_FONT, fontSize=9.2, leading=11,
        textColor=INK, tracking=1.2,
    )
    legend_style = ParagraphStyle(
        'Legend', fontName=BODY_FONT, fontSize=5.5, leading=7,
        textColor=MUTED, alignment=TA_RIGHT,
    )
    detail_heading = Table([
        [paragraph('勤務詳細  /  WORK DETAILS', detail_heading_style),
         paragraph('淡いオレンジ=欠席　緑=振替　青=講習会　赤=推定時間', legend_style)]
    ], colWidths=[90 * mm, 88 * mm])
    detail_heading.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    story = [
        Spacer(1, 5 * mm),
        build_identity_block(sheet, label, period_label),
        Spacer(1, 6 * mm),
        build_metric_grid(sheet),
        Spacer(1, 6 * mm),
        KeepTogether([detail_heading, Spacer(1, 2.2 * mm)]),
        build_detail_table(sheet),
    ]
    doc.build(
        story,
        onFirstPage=lambda canvas, document: draw_page(canvas, document, label, period_label),
        onLaterPages=lambda canvas, document: draw_page(canvas, document, label, period_label),
    )
    return period_prefix, label


def available_path(directory: Path, file_name: str) -> Path:
    candidate = directory / file_name
    if not candidate.exists():
        return candidate
    stem = Path(file_name).stem
    suffix = Path(file_name).suffix
    for index in range(1, 1000):
        candidate = directory / f'{stem} ({index}){suffix}'
        if not candidate.exists():
            return candidate
    raise RuntimeError(f'Could not choose an available name for {file_name}')


def main() -> None:
    if len(sys.argv) < 3:
        raise SystemExit('Usage: build-teacher-pdfs.py <xlsx-path> <output-directory> [qa-directory]')
    workbook_path = Path(sys.argv[1]).resolve()
    output_directory = Path(sys.argv[2]).resolve()
    qa_directory = Path(sys.argv[3]).resolve() if len(sys.argv) >= 4 else Path(tempfile.mkdtemp(prefix='teacher-pdfs-'))
    output_directory.mkdir(parents=True, exist_ok=True)
    qa_directory.mkdir(parents=True, exist_ok=True)
    register_fonts()
    workbook = load_workbook(workbook_path, data_only=False)
    teacher_sheets = [sheet for sheet in workbook.worksheets if sheet.title != '集計一覧']
    if not teacher_sheets:
        raise RuntimeError('講師別シートが見つかりません')

    pdf_files: list[Path] = []
    period_prefix = ''
    for sheet in teacher_sheets:
        sheet_period, label = detect_period(sheet)[0], teacher_label(sheet.title)
        period_prefix = period_prefix or sheet_period
        pdf_path = qa_directory / f'{sheet_period}_{label}.pdf'
        build_teacher_pdf(sheet, pdf_path)
        pdf_files.append(pdf_path)

    zip_path = available_path(output_directory, f'{period_prefix}_講師別PDF.zip')
    with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED, compresslevel=6) as archive:
        for pdf_path in pdf_files:
            archive.write(pdf_path, arcname=pdf_path.name)

    print(json.dumps({
        'pdfCount': len(pdf_files),
        'zipPath': str(zip_path),
        'qaDirectory': str(qa_directory),
        'pdfPaths': [str(path) for path in pdf_files],
    }, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
